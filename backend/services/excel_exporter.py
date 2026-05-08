import re
from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THRESHOLD = 0.1
_DARK     = "252525"
_WHITE    = "FFFFFF"
_RED_BG   = "FDECEA"
_RED_FG   = "E0301E"
_GRAY_BG  = "F0F0F0"
_GRAY_FG  = "808080"
_FONT     = "Malgun Gothic"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold: bool = False, color: Optional[str] = None,
          size: int = 11, italic: bool = False) -> Font:
    kw: Dict[str, Any] = {"name": _FONT, "bold": bold, "size": size, "italic": italic}
    if color:
        kw["color"] = color
    return Font(**kw)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _safe_name(name: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "", name)[:31]


def _rate(prior: Optional[float], current: Optional[float]) -> Optional[float]:
    if not prior or current is None:
        return None
    return (current - prior) / abs(prior)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _build_summary(ws: Any, project: Any, accounts: List[Any]) -> None:
    headers = ["계정과목", "유형", "분석모드", "전기", "당기", "증감액", "증감률",
               "분석기간", "중요성금액", "상태"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _font(bold=True, color=_WHITE)
        c.fill = _fill(_DARK)
        c.alignment = _align(h="center")
        c.border = _border()

    for r, acc in enumerate(accounts, 2):
        comps = acc.components
        prior_t = sum(c.priorAmount or 0 for c in comps)
        cur_t   = sum(c.currentAmount or 0 for c in comps)
        delta   = cur_t - prior_t
        rate    = _rate(prior_t, cur_t)
        mode    = "잔액 자동 필터" if acc.analysisMode == "auto_filter" else "구성요소 직접 정의"
        period  = acc.analysisPeriod
        period_str = f"{period['start']} ~ {period['end']}" if period else ""
        mat = acc.materialityAmount or project.materialityAmount or None

        row_vals = [acc.name, acc.type, mode, prior_t, cur_t, delta,
                    rate, period_str, mat, acc.status]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font()
            c.border = _border()
            if col in (4, 5, 6, 9):
                c.number_format = "#,##0"
                c.alignment = _align(h="right")
            elif col == 7 and val is not None:
                c.number_format = "0.0%"
                c.alignment = _align(h="right")
                if abs(val) > THRESHOLD:
                    c.fill = _fill(_RED_BG)
                    c.font = _font(bold=True, color=_RED_FG)

    col_widths = [20, 8, 20, 15, 15, 15, 10, 25, 12, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ── Account detail sheet ───────────────────────────────────────────────────────

def _build_detail(ws: Any, project: Any, acc: Any) -> None:
    is_auto = acc.analysisMode == "auto_filter"
    mat     = acc.materialityAmount or project.materialityAmount or 0
    comps   = acc.components

    for col_letter, w in [("A", 25), ("B", 50), ("C", 15),
                           ("D", 15), ("E", 15), ("F", 10)]:
        ws.column_dimensions[col_letter].width = w

    r = 1

    # Row 1: Title
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=f"{project.companyName}  분석적 검토 조서")
    c.font = _font(bold=True, size=14)
    ws.row_dimensions[r].height = 28
    r += 2  # blank

    # Meta rows
    cp = project.currentPeriod
    pp = project.priorPeriod
    ap = acc.analysisPeriod or cp

    ws.cell(row=r, column=1, value=f"계정과목: {acc.name}").font = _font()
    ws.cell(row=r, column=4, value=f"유형: {acc.type}").font = _font()
    r += 1
    ws.cell(row=r, column=1,
            value=f"분석기간: {ap.get('start','')} ~ {ap.get('end','')}").font = _font()
    ws.cell(row=r, column=4,
            value=f"전기기간: {pp['start']} ~ {pp['end']}").font = _font()
    r += 1
    mode_label = "잔액 자동 필터" if is_auto else "구성요소 직접 정의"
    ws.cell(row=r, column=1, value=f"분석기준: {mode_label}").font = _font()
    r += 1
    if is_auto and mat:
        ws.cell(row=r, column=1,
                value=f"중요성 금액: {int(mat):,}천원 이상 개별 분석").font = _font()
        r += 1
    r += 1  # blank

    # Section 1: AI Summary
    ws.cell(row=r, column=1, value="1. 증감 총평").font = _font(bold=True)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=acc.aiSummary or "")
    c.font = _font()
    c.alignment = _align(h="left", v="top", wrap=True)
    ws.row_dimensions[r].height = 72
    r += 2  # blank

    # Section 2: Components
    ws.cell(row=r, column=1, value="2. 구성요소별 분석").font = _font(bold=True)
    r += 1

    for col, h in enumerate(["구성요소", "증감 분석 사유", "전기", "당기", "증감액", "증감률"], 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = _font(bold=True, color=_WHITE)
        c.fill = _fill(_DARK)
        c.alignment = _align(h="center")
        c.border = _border()
    r += 1

    individual = [c for c in comps if c.isIndividual]
    aggregate  = [c for c in comps if not c.isIndividual]

    for comp in individual:
        prior = comp.priorAmount or 0
        cur   = comp.currentAmount or 0
        d     = cur - prior
        rate  = _rate(prior, cur)
        for col, val in enumerate([comp.name, comp.aiReason or "", prior, cur, d, rate], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font()
            c.border = _border()
            if col == 2:
                c.alignment = _align(h="left", v="top", wrap=True)
            elif col in (3, 4, 5):
                c.number_format = "#,##0"
                c.alignment = _align(h="right")
            elif col == 6 and val is not None:
                c.number_format = "0.0%"
                c.alignment = _align(h="right")
                if abs(val) > THRESHOLD:
                    c.fill = _fill(_RED_BG)
                    c.font = _font(bold=True, color=_RED_FG)
        ws.row_dimensions[r].height = 36
        r += 1

    if aggregate:
        sp = sum(c.priorAmount or 0 for c in aggregate)
        sc = sum(c.currentAmount or 0 for c in aggregate)
        sd = sc - sp
        sr = _rate(sp, sc)
        for col, val in enumerate(["기타 (합산)", "중요성 미만 항목 합계",
                                    sp, sc, sd, sr], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = _fill(_GRAY_BG)
            c.font = _font(italic=True, color=_GRAY_FG)
            c.border = _border()
            if col in (3, 4, 5):
                c.number_format = "#,##0"
                c.alignment = _align(h="right")
            elif col == 6 and val is not None:
                c.number_format = "0.0%"
                c.alignment = _align(h="right")
        ws.row_dimensions[r].height = 20
        r += 1

    tp  = sum(c.priorAmount or 0 for c in comps)
    tc  = sum(c.currentAmount or 0 for c in comps)
    td  = tc - tp
    tr_ = _rate(tp, tc)
    for col, val in enumerate(["합계", "", tp, tc, td, tr_], 1):
        c = ws.cell(row=r, column=col, value=val)
        c.fill = _fill(_DARK)
        c.font = _font(bold=True, color=_WHITE)
        c.border = _border()
        if col in (3, 4, 5):
            c.number_format = "#,##0"
            c.alignment = _align(h="right")
        elif col == 6 and val is not None:
            c.number_format = "0.0%"
            c.alignment = _align(h="right")
    ws.row_dimensions[r].height = 20
    r += 1

    if is_auto and comps:
        total_abs = sum(abs(c.currentAmount or 0) for c in comps)
        ind_abs   = sum(abs(c.currentAmount or 0) for c in individual)
        pct = (ind_abs / total_abs * 100) if total_abs else 0.0
        c = ws.cell(row=r, column=1, value=f"개별분석 커버리지: {pct:.1f}%")
        c.font = _font(italic=True, color=_GRAY_FG)
        r += 1

    r += 1  # blank

    # Section 3: Auditor Memo
    ws.cell(row=r, column=1, value="3. 감사인 메모").font = _font(bold=True)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=acc.auditorMemo or "")
    c.font = _font()
    c.alignment = _align(h="left", v="top", wrap=True)
    ws.row_dimensions[r].height = 54
    r += 1

    if not is_auto and acc.auditorInstruction:
        r += 1
        ws.cell(row=r, column=1, value="분석 방향 지시").font = _font(bold=True)
        r += 1
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value=acc.auditorInstruction)
        c.font = _font()
        c.alignment = _align(h="left", v="top", wrap=True)
        ws.row_dimensions[r].height = 54


# ── Entry point ───────────────────────────────────────────────────────────────

def build_excel(project: Any, accounts: List[Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "요약"
    _build_summary(ws_sum, project, accounts)

    for acc in accounts:
        ws = wb.create_sheet(title=_safe_name(acc.name))
        _build_detail(ws, project, acc)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
